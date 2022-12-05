import openpyxl
import time

# curTime = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time()))
curTime = time.strftime('%Y-%m-%d', time.localtime(time.time()))
fileName = curTime + '-' + '已检测人员名单'

# 得到所有所有工作簿
xlsList = ['{}.xlsx'.format(name) for name in range(1, 3)]


# 返回一个工作簿的工作表
def getWsList(sh):
    wb = openpyxl.load_workbook(sh)
    wsList = wb.worksheets
    return wsList
    # wb = openpyxl.load_workbook(xlsList[0])
    # wsList = [wb[name] for name in wb.sheetnames]


# 解析工作表中的数据, 获取后写入文件
def getList(ws):
    ws1 = ws
    # for cel in ws1['B']:
    #     print(cel.value)
    with open(f'{fileName}.txt', 'a', encoding='utf-8') as f:
        for i, x in enumerate(list(ws1.columns)[7]):
            if x.fill.fgColor.rgb != '00000000':
                # print(i, x.value)
                nameList = list(ws1.columns)[2]
                numberList = list(ws1.columns)[5]
                f.write(str(nameList[i].value) + ' ' + str(numberList[i].value) + '\n')

# # 解析工作表中的数据, 获取后写入文件
# def getList(ws):
#     ws1 = ws
#     # for cel in ws1['B']:
#     #     print(cel.value)
#     with open(f'{fileName}.xlsx', 'a', encoding='utf-8') as f:
#         re_wb = openpyxl.load_workbook(f'{fileName}.xlsx')
#         re_ws = re_wb['Sheet1']
#         n = 1
#         for i, x in enumerate(list(ws1.columns)[7]):
#             if x.fill.fgColor.rgb != '00000000':
#                 # print(i, x.value)
#                 nameList = list(ws1.columns)[2]
#                 numberList = list(ws1.columns)[5]
#                 # f.write(str(nameList[i].value) + ' ' + str(numberList[i].value) + '\n')
#                 re_ws['A' + str(n)] = namelist[i].value
#                 re_ws['B'+ str(n)] = numberList[i].value
#                 n = n + 1

                
def main():
    for st in xlsList:
        wsList = getWsList(st)
        for ws in wsList:
            getList(ws)


if __name__ == '__main__':
    main()
